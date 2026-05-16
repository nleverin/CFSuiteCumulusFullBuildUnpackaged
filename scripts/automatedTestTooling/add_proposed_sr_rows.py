"""Append rows to the 'Proposed - Service Requests & Case Management'
section of documentation/Automated Testing CFSuite.xlsx, allocating the
next available CFSUITE-SR-NNN target refs.

One-shot helper. Edit NEW_ROWS, run once, commit the spreadsheet.
"""
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font

XLSX = Path("documentation/Automated Testing CFSuite.xlsx")
SHEET = "Automated Tests"
HEADER_ROW = 3
PROPOSED_SR = "Proposed - Service Requests & Case Management"
TARGET_COL_HEADER = "Target Automation Reference"

# 8 new tests: 4 guest scenarios + 4 logged-in repeats
NEW_ROWS = [
    {
        "id": "PROP-059",
        "name": "Guest - Submit case via Request/Location/Details/Complete with map next step",
        "expected": (
            "An unauthenticated (guest) user can submit a service request through the "
            "four-step Request -> Location -> Details -> Complete wizard for a category "
            "whose final response Request Flow has Next Step = Map. The resulting Case "
            "carries the location pin selected on the map."
        ),
        "acceptance": (
            "Case record created with cfsuite1__Request_location__c populated; "
            "cfsuite1__Location__Latitude__s and cfsuite1__Location__Longitude__s set "
            "from the map pin; no community User linked to the case; ContactId resolved "
            "from the matched/new Contact via GuestUserInsertContactHelperClass."
        ),
    },
    {
        "id": "PROP-060",
        "name": "Guest - Submit case via Request/Details/Complete with flow next step",
        "expected": (
            "An unauthenticated (guest) user can submit a service request through the "
            "three-step Request -> Details -> Complete wizard for a category whose final "
            "response Request Flow has Next Step = Flow. The wizard does NOT show the "
            "Location step; the embedded Salesforce Flow captures the request details."
        ),
        "acceptance": (
            "Case record created via the configured Salesforce Flow; flow-collected "
            "fields are mapped onto the Case; no cfsuite1__Request_location__c is "
            "populated; no community User linked; ContactId resolved via guest helpers."
        ),
    },
    {
        "id": "PROP-061",
        "name": "Guest - Submit case with asset layer; verify AssetId captured",
        "expected": (
            "Guest submits via Request -> Location -> Details -> Complete for a category "
            "whose final response Request Flow has Next Step = Map AND an asset layer "
            "configured. Selecting an asset on the map associates the asset with the case."
        ),
        "acceptance": (
            "Case.cfsuite1__AssetId__c is populated with the Id of the asset chosen on "
            "the map layer; map pin location is also captured in "
            "cfsuite1__Request_location__c."
        ),
    },
    {
        "id": "PROP-062",
        "name": "Guest - Submit case with zone layer; verify Zone captured",
        "expected": (
            "Guest submits via Request -> Location -> Details -> Complete for a category "
            "whose final response Request Flow has Next Step = Map AND a zone layer "
            "configured. Selecting a location within a zone associates the zone with the case."
        ),
        "acceptance": (
            "Case.cfsuite1__Zone_GIS__c (and/or cfsuite1__Zones__c) is populated with the "
            "zone name resolved from the map pin position; map pin location also captured "
            "in cfsuite1__Request_location__c."
        ),
    },
    {
        "id": "PROP-063",
        "name": "Logged-in - Submit case via Request/Location/Details/Complete with map next step",
        "expected": (
            "Authenticated community user submits the four-step wizard as in PROP-059. "
            "Case is linked to the user's Account and Contact."
        ),
        "acceptance": (
            "Same outputs as PROP-059, plus: Case.AccountId and Case.ContactId match the "
            "logged-in community user's PersonAccount; CommunityUserInsertCaseHelperClass "
            "executes without error."
        ),
    },
    {
        "id": "PROP-064",
        "name": "Logged-in - Submit case via Request/Details/Complete with flow next step",
        "expected": (
            "Authenticated community user submits the three-step wizard as in PROP-060. "
            "Case is linked to the user's Account and Contact."
        ),
        "acceptance": (
            "Same outputs as PROP-060, plus: Case.AccountId and Case.ContactId match the "
            "logged-in user; cfsuite1__Request_location__c remains empty (no Location step)."
        ),
    },
    {
        "id": "PROP-065",
        "name": "Logged-in - Submit case with asset layer; verify AssetId captured",
        "expected": "Authenticated community user submits as in PROP-061.",
        "acceptance": (
            "Same outputs as PROP-061, plus the Case is linked to the logged-in user's "
            "AccountId / ContactId."
        ),
    },
    {
        "id": "PROP-066",
        "name": "Logged-in - Submit case with zone layer; verify Zone captured",
        "expected": "Authenticated community user submits as in PROP-062.",
        "acceptance": (
            "Same outputs as PROP-062, plus the Case is linked to the logged-in user's "
            "AccountId / ContactId."
        ),
    },
]


def main() -> None:
    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SHEET]

    # Locate the column positions and the last Proposed-SR row.
    cols = {ws.cell(HEADER_ROW, c).value: c for c in range(1, ws.max_column + 1)}
    func_col   = cols["Function"]
    id_col     = cols["Test Case ID"]
    name_col   = cols["Test Case Name/Description"]
    exp_col    = cols["Expected Results"]
    acc_col    = cols["Acceptance Criteria"]
    target_col = cols[TARGET_COL_HEADER]

    last_prop_sr_row = None
    for r in range(4, ws.max_row + 1):
        if str(ws.cell(r, func_col).value).strip() == PROPOSED_SR:
            last_prop_sr_row = r
    assert last_prop_sr_row, "Couldn't find Proposed-SR section"

    # Highest existing SR target ref - new rows continue from there
    sr_pattern = re.compile(r"CFSUITE-SR-(\d+)")
    max_sr_n = 0
    for r in range(4, ws.max_row + 1):
        v = ws.cell(r, target_col).value
        if not v:
            continue
        m = sr_pattern.match(str(v))
        if m:
            max_sr_n = max(max_sr_n, int(m.group(1)))

    # Verify the new PROP IDs aren't already used
    existing_ids = {ws.cell(r, id_col).value for r in range(4, ws.max_row + 1)}
    for new_row in NEW_ROWS:
        assert new_row["id"] not in existing_ids, f"{new_row['id']} already used"

    insert_at = last_prop_sr_row + 1
    print(f"Inserting {len(NEW_ROWS)} rows starting at row {insert_at}")
    print(f"Next SR target ref starts at CFSUITE-SR-{max_sr_n + 1:03d}\n")

    ws.insert_rows(insert_at, amount=len(NEW_ROWS))

    for i, new_row in enumerate(NEW_ROWS):
        r = insert_at + i
        target_ref = f"CFSUITE-SR-{max_sr_n + 1 + i:03d}"
        ws.cell(r, func_col,   value=PROPOSED_SR)
        ws.cell(r, id_col,     value=new_row["id"])
        ws.cell(r, name_col,   value=new_row["name"])
        ws.cell(r, exp_col,    value=new_row["expected"])
        ws.cell(r, acc_col,    value=new_row["acceptance"])
        ws.cell(r, target_col, value=target_ref)
        ws.cell(r, exp_col).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(r, acc_col).alignment = Alignment(wrap_text=True, vertical="top")
        print(f"  row {r}  {new_row['id']:9s}  ->  {target_ref}  {new_row['name'][:60]}")

    wb.save(XLSX)
    print(f"\nWrote {len(NEW_ROWS)} new rows to {XLSX.name}")


if __name__ == "__main__":
    main()
