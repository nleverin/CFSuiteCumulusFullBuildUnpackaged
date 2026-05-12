"""One-shot renumbering of the CFSuite automation:

1. Inserts a new row in `documentation/Automated Testing CFSuite.xlsx` for
   `REQ-006 clone_case` under a new "Additional - Service Requests & Case
   Management" sub-section that sits between the active SR rows and the
   Proposed-SR rows.
2. Inserts a "Target Automation Reference" column to the left of the
   existing "Automation Reference" column.
3. Walks every test row and assigns a sequential `CFSUITE-{Area}-{N}` target
   reference per functional area (ordered: active -> additional -> proposed).
4. Renames each robot file under `robot/tests/requests/` to match its
   target reference and rewrites the file's `[Documentation]` block.
5. Updates the "Automation Reference" column for rows backed by a file.
6. Deletes `robot/tests/requests/REQ-001_create_contact.robot` (boilerplate
   smoke test - not a meaningful CFSuite scenario).

Idempotent: rerunning is a no-op once everything is already named correctly.
"""
import re
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(".").resolve()
XLSX = ROOT / "documentation" / "Automated Testing CFSuite.xlsx"
ROBOT_DIR = ROOT / "robot" / "tests" / "requests"

# --- Configuration -----------------------------------------------------------

AREA_FROM_FUNC = {
    "Authentication & User Management": "AU",
    "Additional - Authentication & User Management": "AU",
    "Proposed - Authentication & User Management": "AU",
    "Service Requests & Case Management": "SR",
    "Additional - Service Requests & Case Management": "SR",
    "Proposed - Service Requests & Case Management": "SR",
    "Communications & Notifications": "CN",
    "Additional - Communications & Notifications": "CN",
    "Proposed - Communications & Notifications": "CN",
    "Councillor Portal": "CP",
    "Additional - Councillor Portal": "CP",
    "Proposed - Councillor Portal": "CP",
    "Configuration & Administration": "CA",
    "Additional - Configuration & Administration": "CA",
    "Proposed - Configuration & Administration": "CA",
}

FUNC_ORDER = {}
for f in AREA_FROM_FUNC:
    if f.startswith("Additional - "):
        FUNC_ORDER[f] = 1
    elif f.startswith("Proposed - "):
        FUNC_ORDER[f] = 2
    else:
        FUNC_ORDER[f] = 0

# Map a CoM-Test row to the existing robot file that automates it.
EXISTING = {
    "CoM-Test-015": "REQ-002_create_case_with_location.robot",
    "CoM-Test-017": "REQ-003_create_case_from_account.robot",
    "CoM-Test-024": "CFSUITE-SR-004_calculate_request_due_date.robot",
    "CoM-Test-025": "REQ-004_create_child_case.robot",
    "CoM-Test-026": "REQ-007_recategorise_case.robot",
    "CoM-Test-027": "CFSUITE-SR-005_flag_request_as_emergency.robot",
    "CoM-Test-028": "CFSUITE-SR-006_due_date_visible_highlights.robot",
    "CoM-Test-048": "REQ-005_assign_case.robot",
    "CoM-Test-049": "CFSUITE-SR-007_link_interested_parties.robot",
    "CoM-Test-052": "CFSUITE-SR-008_track_request_history.robot",
    "CoM-Test-101": "REQ-009_community_follow_case.robot",
    "CoM-Test-106": "REQ-008_customer_portal_login.robot",
}

# Rows we want to add to the spreadsheet that don't correspond to a CoM-Test ID.
EXTRA_ROWS = [
    {
        "function": "Additional - Service Requests & Case Management",
        "id": "REQ-006",
        "name": "Clone Case",
        "expected": (
            "Ability to clone an existing case as a starting point for a "
            "new one. Cloned case carries category and reason fields."
        ),
        "steps": "",
        "data": "",
        "acceptance": "Cloned case opens with the parent's fields pre-filled and can be saved as a new case.",
        "current_file": "REQ-006_clone_case.robot",
    },
]

DELETE_FILES = ["REQ-001_create_contact.robot"]

TARGET_COL_HEADER = "Target Automation Reference"
AUTO_COL_HEADER = "Automation Reference"

# --- Spreadsheet update ------------------------------------------------------

def run_git_mv(src: Path, dst: Path):
    """git mv, falling back to plain rename if the file isn't tracked yet."""
    if src.resolve() == dst.resolve():
        return
    rel_src = src.relative_to(ROOT).as_posix()
    rel_dst = dst.relative_to(ROOT).as_posix()
    result = subprocess.run(
        ["git", "mv", rel_src, rel_dst],
        cwd=ROOT, capture_output=True, text=True,
    )
    if result.returncode != 0:
        # Not tracked yet (e.g. untracked test file) - just rename
        shutil.move(str(src), str(dst))


def find_columns(ws, header_row=3):
    cols = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        cols[str(v).strip()] = c
    return cols


def find_function_block_end(ws, function_value):
    """Return the last row index whose Function == function_value."""
    last = None
    for r in range(4, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == function_value:
            last = r
    return last


def main():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Automated Tests"]
    print(f"Loaded {XLSX.name}: rows=1..{ws.max_row} cols=1..{ws.max_column}")

    cols = find_columns(ws)
    id_col = cols["Test Case ID"]
    func_col = cols["Function"]
    name_col = cols["Test Case Name/Description"]
    expected_col = cols["Expected Results"]
    steps_col = cols["Test Steps"]
    data_col = cols.get("Test Data (If applicable)")
    acc_col = cols["Acceptance Criteria"]
    auto_col_existing = cols.get(AUTO_COL_HEADER)

    # -- 1. Insert extra rows -------------------------------------------------
    # For each extra, find the row immediately after the last active row of
    # its area and insert there.
    inserted_at = {}
    for extra in EXTRA_ROWS:
        # The base function is the area's active heading (strip "Additional - ")
        base_func = extra["function"].replace("Additional - ", "")
        last_active = find_function_block_end(ws, base_func)
        if last_active is None:
            print(f"  WARN: couldn't find base function '{base_func}'")
            continue
        insert_at = last_active + 1
        # If a row was already inserted earlier (from a previous extra in the
        # same area), put this one after it.
        while ws.cell(row=insert_at, column=func_col).value == extra["function"]:
            insert_at += 1
        print(f"  Inserting '{extra['id']}' at row {insert_at} (Function='{extra['function']}')")
        ws.insert_rows(insert_at, amount=1)
        ws.cell(row=insert_at, column=func_col, value=extra["function"])
        ws.cell(row=insert_at, column=id_col, value=extra["id"])
        ws.cell(row=insert_at, column=name_col, value=extra["name"])
        ws.cell(row=insert_at, column=expected_col, value=extra["expected"])
        if extra.get("steps"):
            ws.cell(row=insert_at, column=steps_col, value=extra["steps"])
        if data_col and extra.get("data"):
            ws.cell(row=insert_at, column=data_col, value=extra["data"])
        ws.cell(row=insert_at, column=acc_col, value=extra["acceptance"])
        inserted_at[extra["id"]] = insert_at

    # Re-read columns after inserts (in case max_column shifted - shouldn't, but safe)
    cols = find_columns(ws)
    auto_col_existing = cols.get(AUTO_COL_HEADER)

    # -- 2. Insert the Target Automation Reference column -----------------
    # Target column goes to the LEFT of the existing Automation Reference
    # column. If there's already a Target column, reuse it.
    target_col = cols.get(TARGET_COL_HEADER)
    if target_col is None:
        if auto_col_existing is None:
            # Both missing - append both at the end (target then automation)
            target_col = ws.max_column + 1
            ws.cell(row=3, column=target_col, value=TARGET_COL_HEADER).font = Font(bold=True)
            auto_col = target_col + 1
            ws.cell(row=3, column=auto_col, value=AUTO_COL_HEADER).font = Font(bold=True)
        else:
            ws.insert_cols(auto_col_existing, amount=1)
            target_col = auto_col_existing
            auto_col = auto_col_existing + 1
            ws.cell(row=3, column=target_col, value=TARGET_COL_HEADER).font = Font(bold=True)
        print(f"  Inserted '{TARGET_COL_HEADER}' at column {get_column_letter(target_col)}")
    else:
        auto_col = cols[AUTO_COL_HEADER]
        print(f"  Reusing '{TARGET_COL_HEADER}' at column {get_column_letter(target_col)}")

    # Adjust column widths
    ws.column_dimensions[get_column_letter(target_col)].width = 28
    ws.column_dimensions[get_column_letter(auto_col)].width = 42

    # -- 3. Walk rows, assign sequential targets per area ---------------------
    # Build virtual row list, sort by (area, sub, source_row) -> the same logic
    # as the planner.
    virtual = []
    for r in range(4, ws.max_row + 1):
        func = ws.cell(row=r, column=func_col).value
        test_id = ws.cell(row=r, column=id_col).value
        if not func or not test_id:
            continue
        func_clean = str(func).strip()
        if func_clean not in AREA_FROM_FUNC:
            print(f"  WARN row {r}: unknown function {func!r}")
            continue
        virtual.append({
            "row": r,
            "function": func_clean,
            "id": str(test_id).strip(),
        })

    virtual.sort(key=lambda x: (AREA_FROM_FUNC[x["function"]], FUNC_ORDER[x["function"]], x["row"]))

    counters = {}
    target_for_id = {}
    for v in virtual:
        area = AREA_FROM_FUNC[v["function"]]
        counters[area] = counters.get(area, 0) + 1
        v["target"] = f"CFSUITE-{area}-{counters[area]:03d}"
        target_for_id[v["id"]] = v["target"]

    # Write Target column for each row
    for v in virtual:
        ws.cell(row=v["row"], column=target_col, value=v["target"])

    print("\n  Targets per area:")
    for area, n in sorted(counters.items()):
        print(f"    {area}: {n}")

    # -- 4. Compute rename map for robot files & populate Automation Ref col --
    rename_map = []  # list of (src_file, dst_file, test_id, target)
    for test_id, current_file in EXISTING.items():
        target = target_for_id.get(test_id)
        if target is None:
            print(f"  WARN: {test_id} has no target ref - skipping rename")
            continue
        descriptor = Path(current_file).stem.split("_", 1)[1]
        new_file = f"{target}_{descriptor}.robot"
        rename_map.append((current_file, new_file, test_id, target))

    for extra in EXTRA_ROWS:
        target = target_for_id.get(extra["id"])
        if target is None:
            continue
        descriptor = Path(extra["current_file"]).stem.split("_", 1)[1]
        new_file = f"{target}_{descriptor}.robot"
        rename_map.append((extra["current_file"], new_file, extra["id"], target))

    # Write Automation Reference column
    for v in virtual:
        # Find any rename whose test_id matches this row's id
        match = next((m for m in rename_map if m[2] == v["id"]), None)
        if match is None:
            ws.cell(row=v["row"], column=auto_col, value=None)
            continue
        _, new_file, _, target = match
        rel_path = f"robot/tests/requests/{new_file}"
        cell = ws.cell(row=v["row"], column=auto_col, value=f"{target} ({rel_path})")
        cell.hyperlink = rel_path
        cell.font = Font(color="0563C1", underline="single")
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    # -- 5. Save spreadsheet --------------------------------------------------
    wb.save(XLSX)
    print(f"\n  Saved {XLSX.name}")

    # -- 6. Rename robot files & rewrite their [Documentation] ----------------
    print("\n=== Renaming robot files ===")
    com_for_target = {target: test_id for _, _, test_id, target in rename_map}
    for src_name, dst_name, test_id, target in rename_map:
        src = ROBOT_DIR / src_name
        dst = ROBOT_DIR / dst_name
        if not src.exists() and dst.exists():
            print(f"  {src_name} -> {dst_name}  (already renamed)")
            text = dst.read_text(encoding="utf-8")
        elif not src.exists():
            print(f"  WARN missing source: {src}")
            continue
        else:
            if src.resolve() != dst.resolve():
                run_git_mv(src, dst)
                print(f"  {src_name} -> {dst_name}")
            text = dst.read_text(encoding="utf-8")

        # Update doc references to old name -> new name (and CoM-Test stays).
        old_id = Path(src_name).stem.split("_", 1)[0]  # e.g. CFSUITE-SR-005 or REQ-002
        new_id = target
        new_text = text
        # Replace any '{old_id}' string with new id (literal)
        new_text = new_text.replace(old_id, new_id)
        # Ensure the test's Documentation references its CoM-Test mapping.
        # If no '(CoM-Test-NNN)' appears in the file, add it next to the first
        # CFSUITE-... mention.
        if test_id.startswith("CoM-Test-") and f"({test_id})" not in new_text:
            new_text = new_text.replace(new_id, f"{new_id} ({test_id})", 1)
        if new_text != text:
            dst.write_text(new_text, encoding="utf-8")

    # -- 7. Delete files we no longer want ------------------------------------
    print("\n=== Deleting boilerplate files ===")
    for name in DELETE_FILES:
        p = ROBOT_DIR / name
        if not p.exists():
            print(f"  {name}  (already absent)")
            continue
        result = subprocess.run(
            ["git", "rm", p.relative_to(ROOT).as_posix()],
            cwd=ROOT, capture_output=True, text=True,
        )
        if result.returncode == 0:
            print(f"  git rm {name}")
        else:
            p.unlink()
            print(f"  rm (untracked) {name}")

    print("\nDone.")


if __name__ == "__main__":
    main()
