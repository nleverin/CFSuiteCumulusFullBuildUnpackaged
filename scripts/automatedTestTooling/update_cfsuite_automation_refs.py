"""Refresh the Automation Reference column in
`documentation/Automated Testing CFSuite.xlsx`.

Auto-discovers every robot file under `robot/tests/requests/` named
`CFSUITE-{Area}-{N}_{descriptor}.robot`, matches its `CFSUITE-...` prefix
against the `Target Automation Reference` column in the spreadsheet, and
writes `<target> (<relative path>)` plus a hyperlink into the
`Automation Reference` column.

Idempotent: rerun any time a new CFSUITE-* test file lands. Rows whose
target isn't backed by a file get the Automation Reference cleared.
"""
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font

XLSX = Path("documentation/Automated Testing CFSuite.xlsx")
ROBOT_DIR = Path("robot/tests/requests")
TARGET_HEADER = "Target Automation Reference"
AUTO_HEADER = "Automation Reference"
HEADER_ROW = 3


def main() -> None:
    existing: dict[str, Path] = {}
    for p in sorted(ROBOT_DIR.glob("CFSUITE-*.robot")):
        target = p.stem.split("_", 1)[0]
        if target in existing:
            print(f"  WARN duplicate target {target}: {existing[target]} and {p}")
        existing[target] = p

    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Automated Tests"]

    target_col = auto_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(HEADER_ROW, c).value
        if v == TARGET_HEADER:
            target_col = c
        elif v == AUTO_HEADER:
            auto_col = c
    if not target_col or not auto_col:
        raise SystemExit(
            f"Could not find required columns. target_col={target_col} "
            f"auto_col={auto_col}. Has the renumber script been run?"
        )

    written = cleared = 0
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        target = ws.cell(r, target_col).value
        if not target:
            continue
        target = str(target).strip()
        cell = ws.cell(r, auto_col)
        if target in existing:
            rel = f"robot/tests/requests/{existing[target].name}"
            cell.value = f"{target} ({rel})"
            cell.hyperlink = rel
            cell.font = Font(color="0563C1", underline="single")
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            written += 1
        elif cell.value:
            cell.value = None
            cell.hyperlink = None
            cleared += 1

    wb.save(XLSX)
    print(f"Wrote {written} reference(s), cleared {cleared} stale reference(s) in {XLSX.name}")


if __name__ == "__main__":
    main()
