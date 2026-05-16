"""Append rows to Proposed sections in documentation/Automated Testing CFSuite.xlsx,
sourced from the ESS Knowledge release-note regression test articles
(Releases 3.151 through 3.156).

Each row references the originating ticket and release. SR-area items
go into Proposed - Service Requests & Case Management; AU and CN items
go into their respective Proposed sub-sections.
"""
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment

XLSX = Path("documentation/Automated Testing CFSuite.xlsx")
SHEET = "Automated Tests"
HEADER_ROW = 3
TARGET_COL_HEADER = "Target Automation Reference"

PROPOSED_FUNC = {
    "SR": "Proposed - Service Requests & Case Management",
    "AU": "Proposed - Authentication & User Management",
    "CN": "Proposed - Communications & Notifications",
    "CP": "Proposed - Councillor Portal",
    "CA": "Proposed - Configuration & Administration",
}

# id, area, name, expected, acceptance, source (release + tickets)
NEW_ROWS = [
    # ---- SR area ----
    ("PROP-067", "SR", "Map layer result limits configurable",
     "Custom settings BusLayerMaxResults/Suggestions, BuildingLayerMaxResults/Suggestions, ReserveLayerMaxResults/Suggestions, AddressLayerMaxResults/Suggestions control the max rows returned by the corresponding map layer in the Internal Case Creation view.",
     "Setting each *MaxResults__c to 20 (or 100 for Address) caps the search results at that count. Search suggestions appear after 2 chars for Bus Stops and 3 chars for other layers.",
     "Release 3.156 (47118, 31415)"),
    ("PROP-068", "SR", "My Requests table shows Request Reason column in portal",
     "Logged-in community user viewing the My Requests tile sees a Request Reason column on each row.",
     "Request Reason column is present and shows the case's cfReq_Request_Reason value for each row.",
     "Release 3.156 (34327)"),
    ("PROP-069", "SR", "Asset layer opacity configurable on Map Layers",
     "Opacity setting on a CFSuite Map Layers asset record controls the transparency of the rendered asset overlay on the map.",
     "Setting a lower opacity makes the asset layer more transparent so the underlying base map is visible through it.",
     "Release 3.156 (39932, 41244)"),
    ("PROP-070", "SR", "Create Case button disabled until asset selected - category first",
     "When an asset-layer category is chosen first (then a location), the Create Case button remains disabled until an asset is selected on the map.",
     "Button stays disabled after category/reason/sub-reason selection; enables once an asset is clicked; created case has cfsuite1__AssetId__c populated.",
     "Release 3.156 (46375)"),
    ("PROP-071", "SR", "Create Case button disabled until asset selected - location first",
     "When a location is chosen first (then an asset-layer category), the Create Case button stays disabled until an asset is selected on the map layer.",
     "Reverse order from PROP-070: location selected, then category/reason/sub-reason that displays asset layer; button stays disabled until asset chosen; case carries cfsuite1__AssetId__c.",
     "Release 3.156 (46375)"),
    ("PROP-072", "SR", "Guest routing - 'Optional contact details' + SMS entered",
     "Community Request configured as 'Allow anonymous with optional contact details'. Guest user submits with Notification Method = SMS and a mobile number.",
     "Case is created; Web_Name/Web_Email/Web_Phone fields populated; SMS_Message__c record with Status='Pending' created; warning text from Guest Comms Warning setting visible during the wizard.",
     "Release 3.156 (47242)"),
    ("PROP-073", "SR", "Guest routing - 'Optional contact details' + Email entered",
     "Same Community Request configuration as PROP-072 but guest provides Email + Notification Method = Email.",
     "Case is created; Web_* fields populated; email dispatched to the guest's address; warning visible on contact-details page.",
     "Release 3.156 (47242)"),
    ("PROP-074", "SR", "Guest routing - 'Optional contact details' + No details + No updates",
     "Same configuration as PROP-072 but guest provides no contact details and ticks 'No updates required'.",
     "Submit succeeds; Case is created with Notification Method blank; Description has no guest-user info appended; no notifications sent.",
     "Release 3.156 (47242)"),
    ("PROP-075", "SR", "Guest routing - validation rule SMS-channel mismatch",
     "Same configuration as PROP-072 but guest enters a mobile number while selecting Email as the Notification Method.",
     "No_Phone_And_Notification_Method_Is_SMS / No_Email_And_Notification_Method_Is_Email validation rule fires; Case cannot be saved.",
     "Release 3.156 (47242)"),
    ("PROP-076", "SR", "Guest routing - validation rule Email-channel mismatch",
     "Same configuration as PROP-075 but reverse: guest enters Email while selecting SMS as Notification Method.",
     "Validation rule fires; Case cannot be saved.",
     "Release 3.156 (47242)"),
    ("PROP-077", "SR", "Guest routing - 'Allow anonymous with contact details' mandatory",
     "Community Request configured as 'Allow anonymous with contact details' (NOT optional). Guest submits.",
     "Contact details section is shown with mandatory fields; submitting empty fields shows validation errors; valid submit creates case with Web_* fields populated and Notification Method set; email is received.",
     "Release 3.156 (47242)"),
    ("PROP-078", "SR", "Guest routing - 'Do not allow anonymous'",
     "Community Request configured to disallow anonymous submission.",
     "Guest viewing the category does NOT see a 'Continue as Guest User' button; only the login flow is available.",
     "Release 3.156 (47242) / 3.152 (23704, 26146, 27230)"),
    ("PROP-079", "SR", "Recategorisation due date rebaseline toggle",
     "When 'Enable Recat Due Date Rebaseline' is OFF, recategorising a case to a request flow with a different SLA leaves Milestone_Start_Date unchanged but updates Original_Due_Date based on new SLA from the original start. When ON, both dates rebaseline from the recategorisation time.",
     "Toggle OFF: Milestone_Start_Date unchanged after recat; Original_Due_Date reflects new SLA from original start. Toggle ON: Milestone_Start_Date = recategorisation time; Original_Due_Date = recat time + new SLA.",
     "Release 3.156 (43127)"),
    ("PROP-080", "SR", "AO Map Select All / Clear All filter behaviour",
     "On the AO Map > More Filters > Team and Work Category, the 'All' toggle and 'Clear All' link work consistently with individual flags.",
     "Click 'All': all child flags checked; click 'All' again: all unchecked. Click 'Clear All' link: all individual flags clear.",
     "Release 3.156 (51034)"),
    ("PROP-081", "SR", "Description preserved through case recategorisation",
     "Description entered on a case is retained when the case is recategorised via 'Correct Case Category'.",
     "After recategorisation, Description field on the case still contains the original text.",
     "Release 3.156 (51132)"),
    ("PROP-082", "SR", "Create case via bus-stop layer captures bus-stop address",
     "Internal case create: choose a category/reason that uses the bus-stop layer; select a bus stop on the map; create case.",
     "Case Request_location is populated with the bus stop's address attribute.",
     "Release 3.154 (21079, 33892, 37490)"),
    ("PROP-083", "SR", "Create case via building layer captures building address",
     "Same as PROP-082 but with the building layer.",
     "Case Request_location is populated with the building's address attribute.",
     "Release 3.154 (21079, 33892, 37490) / 3.152 (21079, 23530, 23732, 25327)"),
    ("PROP-084", "SR", "Create case via reserve layer captures reserve address",
     "Same as PROP-082 but with the reserve layer.",
     "Case Request_location is populated with the reserve's address attribute.",
     "Release 3.154 (21079, 33892, 37490)"),
    ("PROP-085", "SR", "Show All Cases Default setting on internal case view",
     "Custom setting 'Show All Cases Default' controls the default state of the Show All Cases toggle in the internal Case Creation view.",
     "Setting = Yes: toggle defaults ON. Setting = No: toggle defaults OFF. Reflects current setting on next page load.",
     "Release 3.154 (24016)"),
    ("PROP-086", "SR", "Description retained through category change on internal case create",
     "On the internal case creation view, entering a Description and then changing the category does not clear the Description.",
     "Description text persists across category/reason/sub-reason changes; persists on the final saved Case.",
     "Release 3.154 (34323)"),
    ("PROP-087", "SR", "Portal case with flow accepts multi-file attachments",
     "On the community portal, creating a case that uses a Flow form supports the configured number of file attachments (per CFSuite Settings).",
     "Attachments setting = 5: portal accepts up to 5 files; each uploaded file is listed; submitted case has all files attached as ContentVersion + ContentDocumentLink.",
     "Release 3.154 (29954, 22612)"),
    ("PROP-088", "SR", "AO Map pin popup, saved filters, manual assignment",
     "Action Officer Map: clicking a pin shows the popup; saved filter loads records; new filter can be created and saved; assign-to flow assigns ownership; list view reflects the active filter.",
     "Pin popup displays; saved filter returns matching cases; new filter persists; case OwnerId updates after assign-to; list view filtered consistently.",
     "Release 3.154 (36041)"),
    ("PROP-089", "SR", "Portal category search suggested list clears when input empty",
     "Customer portal category search bar shows suggested categories as the user types; clearing the input removes the suggestions.",
     "Typing returns suggestions; deleting all characters hides the suggestions list.",
     "Release 3.154 (32651)"),
    ("PROP-090", "SR", "AO Map list view shows Due Date column and sorts by it",
     "AO Map list tab displays a Due Date column; clicking the header sorts by due date asc/desc.",
     "Due Date column visible; sort order changes on header click; rows ordered correctly.",
     "Release 3.152 (17300, 24341, 23471)"),
    ("PROP-091", "SR", "AO Map default view (List/Map) controlled by setting",
     "CFSuite Settings 'AO Map Default view' setting controls whether the AO Map opens in List or Map view on the Home page.",
     "Setting = List: AO Map opens to list view. Setting = Map: AO Map opens to map view.",
     "Release 3.152 (19846)"),
    ("PROP-092", "SR", "Console - clicking case number opens case in new console tab",
     "From a console view such as CFSuite Console, clicking a case number from the home page opens the case in a new console tab (not navigates away).",
     "New console tab opens with the case detail loaded; original tab stays intact.",
     "Release 3.152 (19112)"),
    ("PROP-093", "SR", "Updating Customer Name auto-updates Contact Name on case",
     "On a Case owned by a Person Account: updating the Customer Name (Account lookup) to a Business Account auto-updates the Contact Name field to the related contact.",
     "Contact Name field reflects the new account's primary contact after Customer Name change.",
     "Release 3.152 (30313)"),
    ("PROP-094", "SR", "Edit Location preserves asset Id when pin not moved",
     "On a case created via an asset layer, clicking Edit Location and then Update Location without moving the pin leaves the asset Id intact.",
     "After Edit Location + Update Location with unchanged pin, Case.cfsuite1__AssetId__c is unchanged.",
     "Release 3.152 (23472)"),
    ("PROP-095", "SR", "Additional Details Mandatory portal setting enforces field",
     "CFSuite Settings 'Additional Details Mandatory' set to Yes makes the additional information field required at the end of the request wizard.",
     "Submitting an empty additional-info field shows the validation error; filling and submitting allows the request to proceed to the thank-you page.",
     "Release 3.152 (26144)"),
    ("PROP-096", "SR", "Guest user 'Allow anonymous with no contact details' - no contact form",
     "Community Request configured with 'Allow anonymous with no contact details': guest user submission skips the contact details step entirely.",
     "After location selection, no contact-details section displayed; submit goes directly to thank-you; no notifications dispatched.",
     "Release 3.156 (47242) - Test 6"),
    ("PROP-097", "SR", "Per-flow Cancel button on My Requests opens configured flow in modal",
     "On a Request Flow record, enabling the cancel button and setting My Request Button Flow + My Request Button Name surfaces a button on the My Requests screen for matching cases. Clicking presents the configured flow in a modal.",
     "Button is visible only on cases that match the configured request flow; click opens the flow in a modal; flow runs to completion and reflects on the case.",
     "Release 3.152 (23707)"),
    ("PROP-098", "SR", "Deflection to custom community page with embedded flow",
     "A Community Request record configured to deflect to a URL pointing at a custom community page that hosts a Flow component. Guest user choosing that category is shown the guest-user login first, then the deflection page with the flow.",
     "Guest user login presented before deflection; community page renders the configured Flow component; flow runs as expected.",
     "Release 3.152 (27879)"),

    # ---- AU area: customer merge, business address display, password update ----
    ("PROP-067-AU", "AU", "Merging duplicate customers consolidates property customer links",
     "Two duplicate customer accounts each linked to the same property via Property_Customer records. Merging the customers consolidates the property_customer relationships.",
     "After merge into the customer with the portal account, only one Property_Customer record remains for that property; the count reflects accurately after refresh.",
     "Release 3.154 (26354, 36188)"),
    ("PROP-068-AU", "AU", "Display Business Address New Customer setting controls visibility",
     "Custom setting 'Display Business Address New Customer' controls whether the address section is visible when creating a new Business customer.",
     "Setting = Yes: address section visible. Setting = No: address section hidden.",
     "Release 3.154 (40339)"),
    ("PROP-069-AU", "AU", "Update password via My Profile - forgot password link",
     "Logged-in community user navigates to My Profile > Account and password > Update Password; clicking the forgotten password link sends a reset email; the link works to reset the password.",
     "Reset email received; following the link allows password change; user can log in with new password.",
     "Release 3.154 (36043) / 3.151"),
    ("PROP-070-AU", "AU", "Update password via My Profile - current/new/confirm",
     "Logged-in community user changes their password using the Current/New/Confirm fields on the My Profile screen.",
     "Submitting valid current + matching new/confirm fields shows the success message and returns to My Profile; mismatched or wrong current password shows the appropriate error.",
     "Release 3.154 (36043)"),

    # ---- CN area: My Local Gov inbox + folder management ----
    ("PROP-067-CN", "CN", "My Local Gov inbox returns and marks messages read",
     "On the My Local Gov portal, navigating to the Inbox shows existing messages with correct unread state; clicking a message marks it read both in the portal and the internal view.",
     "Messages returned in Inbox; unread indicator matches IsRead on the backing record; click toggles to read state.",
     "Release 3.154 (36042)"),
    ("PROP-068-CN", "CN", "My Local Gov create folder and move messages between folders",
     "User creates a new folder under Inbox; drags a message from Inbox to the new folder; the message moves between folders correctly.",
     "New folder is created and visible; message moves out of Inbox into the new folder; both views reflect the move.",
     "Release 3.154 (36042)"),
]


def main() -> None:
    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SHEET]

    cols = {ws.cell(HEADER_ROW, c).value: c for c in range(1, ws.max_column + 1)}
    func_col   = cols["Function"]
    id_col     = cols["Test Case ID"]
    name_col   = cols["Test Case Name/Description"]
    exp_col    = cols["Expected Results"]
    acc_col    = cols["Acceptance Criteria"]
    target_col = cols[TARGET_COL_HEADER]

    # Find the last row for each Proposed-* sub-section and the highest
    # target number in each area
    last_proposed_row = {}
    max_target_n = {}
    target_pattern = re.compile(r"CFSUITE-([A-Z]{2})-(\d+)")
    for r in range(4, ws.max_row + 1):
        func = ws.cell(r, func_col).value
        v = ws.cell(r, target_col).value
        if v:
            m = target_pattern.match(str(v))
            if m:
                area = m.group(1)
                max_target_n[area] = max(max_target_n.get(area, 0), int(m.group(2)))
        for area, proposed_label in PROPOSED_FUNC.items():
            if str(func).strip() == proposed_label:
                last_proposed_row[area] = r

    # Idempotent: drop any rows whose Test Case ID is already in the sheet
    existing_ids = {ws.cell(r, id_col).value for r in range(4, ws.max_row + 1)}
    to_add = [row for row in NEW_ROWS if row[0] not in existing_ids]
    skipped = len(NEW_ROWS) - len(to_add)
    if skipped:
        print(f"Skipping {skipped} row(s) already present by Test Case ID.\n")
    if not to_add:
        print("Nothing to add.")
        return

    # Group inserts by area so we insert them contiguously in each section
    by_area = {}
    for row in to_add:
        by_area.setdefault(row[1], []).append(row)

    # Process areas from highest-row to lowest so earlier insertions don't shift later target rows
    processed_areas = sorted(by_area.keys(), key=lambda a: -last_proposed_row[a])

    next_target = {a: max_target_n[a] for a in PROPOSED_FUNC}

    for area in processed_areas:
        rows = by_area[area]
        insert_at = last_proposed_row[area] + 1
        print(f"\n== {area}: inserting {len(rows)} rows at row {insert_at} ==")
        ws.insert_rows(insert_at, amount=len(rows))
        for i, (rid, _, name, expected, acceptance, source) in enumerate(rows):
            r = insert_at + i
            next_target[area] += 1
            target_ref = f"CFSUITE-{area}-{next_target[area]:03d}"
            ws.cell(r, func_col,   value=PROPOSED_FUNC[area])
            ws.cell(r, id_col,     value=rid)
            ws.cell(r, name_col,   value=name)
            ws.cell(r, exp_col,    value=f"{expected}\n\nSource: {source}")
            ws.cell(r, acc_col,    value=acceptance)
            ws.cell(r, target_col, value=target_ref)
            for c in (exp_col, acc_col, name_col):
                ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
            print(f"  row {r}  {rid:14s}  ->  {target_ref}  {name[:60]}")

    wb.save(XLSX)
    print(f"\nWrote {len(NEW_ROWS)} new rows to {XLSX.name}")


if __name__ == "__main__":
    main()
