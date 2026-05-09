"""Add/update an 'Automation Reference' column in the CFSuite test spreadsheet
so each automated row points at the CFSUITE-{Area}-{N} test that covers it
(and the relative path of the .robot file).

Idempotent: re-running updates existing values rather than duplicating columns.
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment

XLSX = Path("documentation/Automated Testing CFSuite.xlsx")
SHEET = "Automated Tests"
HEADER_LABEL = "Automation Reference"

# Mapping: CoM-Test-XXX -> (CFSUITE-SR-NNN, relative robot file path)
MAPPING = {
    "CoM-Test-024": ("CFSUITE-SR-004", "robot/tests/requests/CFSUITE-SR-004_calculate_request_due_date.robot"),
    "CoM-Test-027": ("CFSUITE-SR-005", "robot/tests/requests/CFSUITE-SR-005_flag_request_as_emergency.robot"),
    "CoM-Test-028": ("CFSUITE-SR-006", "robot/tests/requests/CFSUITE-SR-006_due_date_visible_highlights.robot"),
    "CoM-Test-049": ("CFSUITE-SR-007", "robot/tests/requests/CFSUITE-SR-007_link_interested_parties.robot"),
    "CoM-Test-052": ("CFSUITE-SR-008", "robot/tests/requests/CFSUITE-SR-008_track_request_history.robot"),
}

wb = openpyxl.load_workbook(XLSX)
ws = wb[SHEET]

# Find the row that holds the Test Case ID header
header_row_idx = None
for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
    if any(c and "Test Case ID" in str(c) for c in row):
        header_row_idx = i
        break
assert header_row_idx is not None, "Couldn't find header row containing 'Test Case ID'"

# Find ID column and existing automation-reference column (if any)
id_col = None
ref_col = None
last_col = ws.max_column
for c in range(1, last_col + 1):
    val = ws.cell(row=header_row_idx, column=c).value
    if val and "Test Case ID" in str(val):
        id_col = c
    if val and HEADER_LABEL in str(val):
        ref_col = c

assert id_col is not None
if ref_col is None:
    ref_col = last_col + 1
    cell = ws.cell(row=header_row_idx, column=ref_col, value=HEADER_LABEL)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions[cell.column_letter].width = 40
    print(f"Added new column at index {ref_col} ({cell.column_letter})")
else:
    print(f"Reusing existing column at index {ref_col}")

updates = 0
for r in range(header_row_idx + 1, ws.max_row + 1):
    test_id = ws.cell(row=r, column=id_col).value
    if not test_id:
        continue
    test_id = str(test_id).strip()
    if test_id not in MAPPING:
        continue
    code, path = MAPPING[test_id]
    cell = ws.cell(row=r, column=ref_col)
    cell.value = f"{code} ({path})"
    cell.hyperlink = path
    cell.font = Font(color="0563C1", underline="single")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    updates += 1
    print(f"  row {r:3d}: {test_id}  ->  {code}")

wb.save(XLSX)
print(f"\nWrote {updates} reference(s) to {XLSX}")
