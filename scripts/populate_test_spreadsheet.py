"""Populate the Automated Testing spreadsheet with test case data."""
import openpyxl
from openpyxl.styles import Alignment

# Read the original spreadsheet
orig_wb = openpyxl.load_workbook(
    "C:/Users/hughw/Downloads/Rates Portal - Master Test Plan Automated Testing.xlsx",
    data_only=True,
)

original_tests = {}
tabs = [
    "3 Create Account",
    "3a. Account Establishment",
    "3b. New Rates Open Cases",
    "3d. Payment History",
    "3e. Property Details",
    "3f. Rates Notices",
    "3g. Making Payments",
    "3h. Refunds and Transfers",
]

for tab_name in tabs:
    ws = orig_wb[tab_name]
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=6, values_only=True):
        function, test_id, description, expected, steps, data = row
        if test_id and test_id not in ("Test Case ID",):
            original_tests[test_id] = {
                "function": function or "",
                "id": test_id,
                "description": description or "",
                "expected": expected or "",
                "steps": steps or "",
                "data": data or "",
            }

# Setup verification tests
setup_tests = [
    {
        "function": "Org Setup",
        "id": "RATES-000a",
        "description": "Verify Admin Has Role",
        "expected": "Admin user has a UserRole assigned (required for community user creation)",
        "steps": (
            "1. Query the current admin user\n"
            "2. Check UserRoleId is not null\n"
            "3. If null, fail with instructions to assign role manually in Setup > Users"
        ),
        "env_data": (
            "Manual setup required:\n"
            "- Assign any Role to admin user in Setup > Users > Edit"
        ),
    },
    {
        "function": "Org Setup",
        "id": "RATES-000b",
        "description": "Verify Community User For Hammy Hampster",
        "expected": "An active community user exists for the Hammy Hampster person account",
        "steps": (
            "1. Query Hammy Hampster account for PersonContactId\n"
            "2. Query User where ContactId matches and IsActive=true\n"
            "3. If no user found, fail with instructions to enable customer user manually"
        ),
        "env_data": (
            "Manual setup required:\n"
            "- Account > Hammy Hampster > dropdown > Enable Customer User\n"
            "- Select License: Customer Community Login\n"
            "- Select Profile: cfsuite community login user"
        ),
    },
    {
        "function": "Org Setup",
        "id": "RATES-000c",
        "description": "Populate Property Rates Data",
        "expected": "All properties have balance data populated (quarterly balances, due dates, financial year)",
        "steps": (
            "1. Query all Property__c records\n"
            "2. For each property without a Full_year_balance__c, set:\n"
            "   - Full_year_balance = $2400\n"
            "   - Financial_Year = 2025\n"
            "   - Quarter 1-4 balances = $600 each\n"
            "   - Quarter 1-4 due dates\n"
            "   - Name_on_Rates_Notice = Hammy Hampster"
        ),
        "env_data": (
            "Properties in dev_namespaced_1 (already had balances):\n"
            "- 37 Goldberry Street: $300.00\n"
            "- 1 Cadisfly Alley: $4,000,000,000.00\n"
            "- 32 Courage Square: $100.00\n"
            "- 1 Humbdinger avenue: -$200.00\n"
            "- 32 Develish Court: $100.00"
        ),
    },
]

# Implemented Account Establishment tests mapped to original IDs
implemented_tests = [
    {
        "orig_id": "CRM_AM01",
        "robot_file": "RATES-010_existing_user_sees_properties.robot",
        "env_data": (
            "Person Account: Hammy Hampster (Authority_ID: 302300)\n"
            "Community User: hammy.hampster@test.example.com\n"
            "Profile: cfsuite community login user\n\n"
            "Properties (Active Primary - should be visible):\n"
            "- 37 Goldberry Street (A) | Balance: $300.00\n"
            "- 1 Cadisfly Alley (A) | Balance: $4,000,000,000.00\n"
            "- 32 Develish Court (A) | Balance: $100.00\n\n"
            "Property Customers (non-Primary - should NOT show):\n"
            "- 32 Courage Square (A) | Secondary | Active\n"
            "- 1 Humbdinger avenue (A) | Previous Primary | Active\n\n"
            "Org Setup Required:\n"
            "- Admin user must have a Role assigned\n"
            "- Community must be Published\n"
            "- Hammy Hampster must have an active community user"
        ),
    },
    {
        "orig_id": "CRM_AM02",
        "robot_file": "RATES-011_cannot_see_deactivated_properties.robot",
        "env_data": (
            "Person Account: Hammy Hampster\n"
            "Community User: hammy.hampster@test.example.com\n\n"
            "Test deactivates one Active Primary Property Customer record\n"
            "via API (sets Relationship_Status__c = Inactive), then\n"
            "verifies the property is no longer visible in the portal.\n\n"
            "Property used: First Active Primary property found\n"
            "(e.g. 37 Goldberry Street, PC ID: a0cRu000007j1OvIAI)\n\n"
            "Cleanup: Test reactivates the property customer after verification."
        ),
    },
    {
        "orig_id": "CRM_AM03",
        "robot_file": "RATES-012_billing_status_visible.robot",
        "env_data": (
            "Person Account: Hammy Hampster\n"
            "Community User: hammy.hampster@test.example.com\n\n"
            "Properties must have balance data populated:\n"
            "- Full_year_balance__c (e.g. $300.00)\n"
            "- Quarter_1-4_Balance__c\n"
            "- Quarter_1-4_Due_Date__c\n"
            "- Financial_Year__c\n\n"
            "Test verifies dollar amounts are displayed on\n"
            "My Linked Properties page and navigates to\n"
            "Manage Property summary page."
        ),
    },
    {
        "orig_id": "CRM_AM04",
        "robot_file": "RATES-013_staff_sees_debt_collection.robot",
        "env_data": (
            "Staff user: Admin (test-je8oob5eor8v@example.com)\n\n"
            "Test sets Memo__c field on a Property to\n"
            "'99 - Debt Collection' via API, then navigates\n"
            "to the Property record in Salesforce (staff view)\n"
            "and verifies the memo is visible.\n\n"
            "Memo__c is a restricted picklist. Valid values:\n"
            "- 99 - Debt Collection\n\n"
            "Property used: First property found\n"
            "(e.g. 37 Goldberry Street, ID: a0dRu00000HzD51IAF)"
        ),
    },
    {
        "orig_id": "CRM_AM05",
        "robot_file": "RATES-014_staff_removes_portal_access.robot",
        "env_data": (
            "Staff user: Admin\n"
            "Person Account: Hammy Hampster\n"
            "Community User: hammy.hampster@test.example.com\n\n"
            "Test navigates to Property Customer record in SF,\n"
            "changes Relationship_Status__c from Active to Inactive\n"
            "(via UI edit button or API fallback), then logs into\n"
            "the portal to verify the property is hidden.\n\n"
            "Property used: First Active Primary property found\n"
            "(e.g. 37 Goldberry Street, PC ID: a0cRu000007j1OvIAI)\n\n"
            "Cleanup: Test reactivates the property customer after verification."
        ),
    },
    {
        "orig_id": "CRM_AM06",
        "robot_file": "RATES-015_cannot_add_properties.robot",
        "env_data": (
            "Person Account: Hammy Hampster\n"
            "Community User: hammy.hampster@test.example.com\n\n"
            "CFSuiteRatesSettings__c:\n"
            "- Allow_Agent_Property_Add__c = No\n\n"
            "Test verifies 'Add Property' button is NOT visible\n"
            "on the My Linked Properties page when the setting is off."
        ),
    },
    {
        "orig_id": "CRM_VAM06a",
        "robot_file": "RATES-016_can_add_properties.robot",
        "env_data": (
            "Person Account: Hammy Hampster\n"
            "Community User: hammy.hampster@test.example.com\n\n"
            "CFSuiteRatesSettings__c:\n"
            "- Allow_Agent_Property_Add__c = Yes (set by test)\n\n"
            "Test temporarily sets the setting to Yes, then verifies\n"
            "'Add Property' button IS visible on My Linked Properties.\n\n"
            "Cleanup: Test restores original setting value (No) after verification."
        ),
    },
]

# Payment History tests
payment_history_tests = [
    {
        "orig_id": "CRM_PH01",
        "robot_file": "RATES-019_view_payment_history.robot",
        "env_data": (
            "Person Account: Hammy Hampster\n"
            "Community User: hammy.hampster@test.example.com\n\n"
            "Pre-requisite: Named Credential 'Payment_Transaction_xxx'\n"
            "must be configured with a valid endpoint for payment\n"
            "history API (stub or live).\n\n"
            "Test navigates to property summary, clicks Payments tab,\n"
            "and verifies payment transaction records are displayed.\n\n"
            "Property: 37 Goldberry Street (A)"
        ),
    },
    {
        "orig_id": "CRM_PH02",
        "robot_file": "RATES-020_payment_delay_disclaimer.robot",
        "env_data": (
            "Person Account: Hammy Hampster\n"
            "Community User: hammy.hampster@test.example.com\n\n"
            "Test navigates to property summary, clicks Payments tab,\n"
            "and verifies a disclaimer about payment processing delay\n"
            "(up to 3 days) is visible on the page.\n\n"
            "Property: 37 Goldberry Street (A)"
        ),
    },
    {
        "orig_id": "CRM_PH03",
        "robot_file": "RATES-021_check_rates_balances.robot",
        "env_data": (
            "Person Account: Hammy Hampster\n"
            "Community User: hammy.hampster@test.example.com\n\n"
            "Test navigates to My Linked Properties page and verifies\n"
            "multiple properties are listed with dollar balance amounts.\n\n"
            "Properties with balances:\n"
            "- 37 Goldberry Street: $300.00\n"
            "- 1 Cadisfly Alley: $4,000,000,000.00\n"
            "- 32 Develish Court: $100.00"
        ),
    },
]

# New Rates Open Cases tests
new_rates_cases_tests = [
    {
        "orig_id": "CRM_RC01",
        "robot_file": "RATES-017_short_term_arrangement.robot",
        "env_data": (
            "Person Account: Hammy Hampster\n"
            "Community User: hammy.hampster@test.example.com\n\n"
            "Portal URL: /s/paymentextension\n"
            "LWC Component: cfsuite1:paymentPlanorExtensionLWC\n\n"
            "Test navigates to property summary to check for payment\n"
            "arrangement section, then goes to the payment extension\n"
            "page and verifies short term arrangement options exist\n"
            "with fine/interest warnings."
        ),
    },
    {
        "orig_id": "CRM_RC02",
        "robot_file": "RATES-018_current_fy_notice_deflection.robot",
        "env_data": (
            "Person Account: Hammy Hampster\n"
            "Community User: hammy.hampster@test.example.com\n\n"
            "Pre-requisite: Named Credential 'Rates_Notice_xxx'\n"
            "must be configured.\n\n"
            "Test navigates to property summary, clicks Rates Notices\n"
            "tab, verifies current FY notices are available as PDF\n"
            "so the customer doesn't need to call for them."
        ),
    },
]

# Payment Plan and Extension tests
payment_plan_tests = [
    {
        "orig_id": "CRM_PH04",
        "robot_file": "RATES-022_request_payment_plan.robot",
        "env_data": (
            "Person Account: Hammy Hampster\n"
            "Community User: hammy.hampster@test.example.com\n\n"
            "Portal URL: /s/paymentextension\n\n"
            "Test navigates to payment extension page and verifies\n"
            "payment plan request option exists with calculator\n"
            "elements (frequency, start date, amounts)."
        ),
    },
    {
        "orig_id": "CRM_PH06",
        "robot_file": "RATES-023_underpayment_plan_alert.robot",
        "env_data": (
            "Person Account: Hammy Hampster\n"
            "Community User: hammy.hampster@test.example.com\n\n"
            "Portal URL: /s/paymentextension\n\n"
            "Test navigates to payment plan page and checks for\n"
            "alert messaging when payment amount is less than\n"
            "required (staff review needed)."
        ),
    },
    {
        "orig_id": "CRM_PH09",
        "robot_file": "RATES-026_arrears_blocks_payment_plan.robot",
        "env_data": (
            "Person Account: Hammy Hampster\n"
            "Community User: hammy.hampster@test.example.com\n\n"
            "Portal URL: /s/paymentextension\n\n"
            "Test navigates to payment plan page and checks for\n"
            "arrears/debt collection messaging that blocks plan setup.\n\n"
            "Note: Property may need to be in arrears for full validation.\n"
            "CFSuiteRatesSettings:\n"
            "- Minimum_Arrears_Amount__c defines the threshold"
        ),
    },
    {
        "orig_id": "CRM_PH10",
        "robot_file": "RATES-027_short_term_less_than_5_days.robot",
        "env_data": (
            "Person Account: Hammy Hampster\n"
            "Community User: hammy.hampster@test.example.com\n\n"
            "Portal URL: /s/paymentextension\n\n"
            "Test navigates to extension page, clicks short term\n"
            "arrangement, verifies <5 days option and reason field."
        ),
    },
    {
        "orig_id": "CRM_PH11",
        "robot_file": "RATES-028_short_term_more_than_5_days.robot",
        "env_data": (
            "Person Account: Hammy Hampster\n"
            "Community User: hammy.hampster@test.example.com\n\n"
            "Portal URL: /s/paymentextension\n\n"
            "Test navigates to extension page, clicks short term\n"
            "arrangement, verifies >5 days option with date selection\n"
            "and staff review messaging."
        ),
    },
    {
        "orig_id": "CRM_PH13",
        "robot_file": "RATES-030_customer_sees_approved_arrangement.robot",
        "env_data": (
            "Person Account: Hammy Hampster\n"
            "Community User: hammy.hampster@test.example.com\n\n"
            "Test navigates to property summary and checks for\n"
            "arrangement status display. Also clicks Property Requests\n"
            "tab to verify case visibility.\n\n"
            "Note: An approved arrangement Case must exist for\n"
            "full validation."
        ),
    },
    {
        "orig_id": "CRM_PH15",
        "robot_file": "RATES-030_customer_sees_approved_arrangement.robot",
        "env_data": (
            "Person Account: Hammy Hampster\n"
            "Community User: hammy.hampster@test.example.com\n\n"
            "Same test file as CRM_PH13 (second test case).\n"
            "Verifies customer can see approved <5 day arrangement\n"
            "on the property summary page."
        ),
    },
]

# Making Payments tests
making_payments_tests = [
    {
        "orig_id": "CRM_SP08",
        "robot_file": "RATES-050_direct_debit_deflection.robot",
        "env_data": (
            "Person Account: Hammy Hampster\n"
            "Community User: hammy.hampster@test.example.com\n\n"
            "Test checks property summary and payment extension page\n"
            "for direct debit references and links to the DD online form.\n\n"
            "Portal URL: /s/paymentextension"
        ),
    },
]

# Rates Notices tests
rates_notices_tests = [
    {
        "orig_id": "CRM_RN01",
        "robot_file": "RATES-042_view_rates_notices.robot",
        "env_data": (
            "Person Account: Hammy Hampster\n"
            "Community User: hammy.hampster@test.example.com\n\n"
            "Pre-requisite: Named Credential 'Rates_Notice_xxx'\n"
            "must be configured with a valid endpoint for rates\n"
            "notices API (stub or live).\n\n"
            "Test checks:\n"
            "1. My Linked Properties page has a rates notice download link\n"
            "2. Manage property > Rates Notices tab shows notice content\n"
            "3. Notice list/table is visible with quarter/year references\n"
            "4. PDF download links are available\n\n"
            "Property: 37 Goldberry Street (A)\n"
            "Portal tabs: Summary | Property Details | Rates Notices | Payments | Property Requests"
        ),
    },
]

# Refunds and Transfers tests
refund_transfer_tests = [
    {
        "orig_id": "CRM_REF01",
        "robot_file": "RATES-051_request_refund.robot",
        "env_data": (
            "Person Account: Hammy Hampster\n"
            "Community User: hammy.hampster@test.example.com\n\n"
            "Portal URL: /s/refund-credit\n"
            "LWC Component: cfsuite1:refundCreditLWC\n\n"
            "Test navigates to the Refund Credit Balance page and\n"
            "verifies the refund form loads with BSB/Account fields.\n\n"
            "CFSuiteRatesSettings:\n"
            "- Refund_Credit_Balance__c = 'Rates - Refund Credit Balance'"
        ),
    },
    {
        "orig_id": "CRM_REF03",
        "robot_file": "RATES-053_request_reverse_payment.robot",
        "env_data": (
            "Person Account: Hammy Hampster\n"
            "Community User: hammy.hampster@test.example.com\n\n"
            "Portal URL: /s/payment-transfer-or-refund\n"
            "LWC Component: cfsuite1:paymentTransferOrRefundLWC\n\n"
            "Test navigates to the Payment Transfer or Refund page\n"
            "and verifies reverse payment options are available.\n\n"
            "CFSuiteRatesSettings:\n"
            "- Reverse_Payment_Bank__c = 'Rates - Reverse Payment (Bank)'"
        ),
    },
    {
        "orig_id": "CRM_REF05",
        "robot_file": "RATES-055_request_transfer_payment.robot",
        "env_data": (
            "Person Account: Hammy Hampster\n"
            "Community User: hammy.hampster@test.example.com\n\n"
            "Portal URL: /s/payment-transfer-or-refund\n"
            "LWC Component: cfsuite1:paymentTransferOrRefundLWC\n\n"
            "Test navigates to the Payment Transfer or Refund page\n"
            "and verifies transfer payment options are available.\n\n"
            "CFSuiteRatesSettings:\n"
            "- Transfer_Payment__c = 'Rates - Transfer Payment'"
        ),
    },
    {
        "orig_id": "CRM_REF07",
        "robot_file": "RATES-057_customer_checks_transfer.robot",
        "env_data": (
            "Person Account: Hammy Hampster\n"
            "Community User: hammy.hampster@test.example.com\n\n"
            "Test navigates to property summary, clicks Property Requests\n"
            "tab, and verifies the request/case information is visible.\n\n"
            "Portal tabs: Summary | Property Details | Rates Notices |\n"
            "Payments | Property Requests\n\n"
            "Property: 37 Goldberry Street (A)"
        ),
    },
]

# Property Details tests
property_details_tests = [
    {
        "orig_id": "CRM_PD01",
        "robot_file": "RATES-037_view_property_details.robot",
        "env_data": (
            "Person Account: Hammy Hampster\n"
            "Community User: hammy.hampster@test.example.com\n\n"
            "Test navigates to My Linked Properties, clicks\n"
            "Manage property for the first property, verifies\n"
            "Assessment Number is visible on the Summary tab,\n"
            "then clicks Property Details tab to check capital value.\n\n"
            "Property: 37 Goldberry Street (A)\n"
            "- Assessment Number: 113256\n"
            "- Council Value: $1,000,000.00\n"
            "- Valuation Number: 100006\n\n"
            "Portal tabs: Summary | Property Details | Rates Notices | Payments | Property Requests"
        ),
    },
    {
        "orig_id": "CRM_PD02",
        "robot_file": "RATES-038_contest_valuation_link.robot",
        "env_data": (
            "Person Account: Hammy Hampster\n"
            "Community User: hammy.hampster@test.example.com\n\n"
            "Test navigates to property summary, clicks\n"
            "Property Details tab, and verifies a link/reference\n"
            "to the Office of the Valuer-General exists for\n"
            "contesting the property valuation.\n\n"
            "Property: 37 Goldberry Street (A)"
        ),
    },
    {
        "orig_id": "CRM_PD05",
        "robot_file": "RATES-041_customer_sees_capital_value.robot",
        "env_data": (
            "Person Account: Hammy Hampster\n"
            "Community User: hammy.hampster@test.example.com\n\n"
            "Test queries the Council_Value__c from the API,\n"
            "then navigates to the property summary, clicks\n"
            "Property Details tab, and verifies the displayed\n"
            "capital value matches the API value.\n\n"
            "Property: 37 Goldberry Street (A)\n"
            "- Council_Value__c: $1,000,000.00\n"
            "- Displayed as: $1,000,000"
        ),
    },
]


# EPT cross-reference tests (from MCC Rates Portal requirements)
ept_tests = [
    {
        "orig_id": "Enno-Rates-03",
        "robot_file": "RATES-060_secondary_sees_property.robot",
        "env_data": (
            "Person Account: Hilda Hampster\n"
            "Community User: hilda@scrat.dev1\n"
            "Relationship: Secondary on 37 Goldberry Street\n\n"
            "Test logs in as Hilda (secondary customer) and verifies\n"
            "she can see the linked property on My Properties & Rates."
        ),
        "manual_entry": {
            "function": "Secondary Relationships",
            "description": "Secondary Customer can see linked property in portal",
            "expected": "Secondary customer can view properties they are linked to as Secondary",
            "steps": "1. Log in as Hilda Hampster (secondary customer) via admin Login as Experience User\n2. Navigate to My Properties & Rates\n3. Verify 37 Goldberry Street is visible",
        },
    },
    {
        "orig_id": "Enno-Rates-04",
        "robot_file": "RATES-061_agent_sees_property.robot",
        "env_data": (
            "Person Account: Horatio Hampster\n"
            "Community User: horatio@scrat.dev1\n"
            "Relationship: Agent on 1 Cadisfly Alley\n\n"
            "Test logs in as Horatio (agent customer) and verifies\n"
            "he can see the linked property on My Properties & Rates."
        ),
        "manual_entry": {
            "function": "Agent Relationships",
            "description": "Agent customer can see linked property in portal",
            "expected": "Agent customer can view properties they manage",
            "steps": "1. Log in as Horatio Hampster (agent) via admin Login as Experience User\n2. Navigate to My Properties & Rates\n3. Verify 1 Cadisfly Alley is visible",
        },
    },
    {
        "orig_id": "Enno-Rates-09",
        "robot_file": "RATES-062_view_managing_agent.robot",
        "env_data": (
            "Person Account: Hammy Hampster (Primary)\n"
            "Agent: Horatio Hampster on 1 Cadisfly Alley\n\n"
            "Test logs in as Hammy, navigates to 1 Cadisfly Alley\n"
            "property summary, clicks Property Details tab and\n"
            "checks for agent information."
        ),
        "manual_entry": {
            "function": "Agent Relationships",
            "description": "Primary customer can view managing agent on property",
            "expected": "Primary customer can see agent linked to their property",
            "steps": "1. Log in as Hammy Hampster via admin Login as Experience User\n2. Navigate to My Properties & Rates\n3. Click Manage property for 1 Cadisfly Alley\n4. Click Property Details tab\n5. Verify agent information is visible",
        },
    },
    {
        "orig_id": "Enno-Rates-25",
        "robot_file": "RATES-063_view_previous_properties.robot",
        "env_data": (
            "Person Account: Hammy Hampster\n"
            "Previous Primary on: 1 Humbdinger avenue\n"
            "Portal URL: /s/previousproperties\n\n"
            "Test navigates to Previous Properties page and verifies\n"
            "previously owned property is listed."
        ),
        "manual_entry": {
            "function": "Previous Properties",
            "description": "Customer can view previously owned properties",
            "expected": "Customer can see properties they previously owned",
            "steps": "1. Log in as Hammy Hampster\n2. Navigate to /s/previousproperties\n3. Verify 1 Humbdinger avenue is listed as a previous property",
        },
    },
    {
        "orig_id": "Enno-Rates-26",
        "robot_file": "RATES-063_view_previous_properties.robot",
        "env_data": (
            "Person Account: Hammy Hampster\n"
            "Current properties: 37 Goldberry Street, 1 Cadisfly Alley, 32 Develish Court\n"
            "Previous property: 1 Humbdinger avenue\n\n"
            "Test verifies both current properties (My Linked Properties)\n"
            "and previous properties (/s/previousproperties) are accessible."
        ),
        "manual_entry": {
            "function": "Previous Properties",
            "description": "Customer can see both current and previous properties",
            "expected": "Customer can view current properties and previously owned properties in separate portal views",
            "steps": "1. Log in as Hammy Hampster\n2. Navigate to My Linked Properties - verify current properties shown\n3. Navigate to Previous Properties - verify previous properties shown",
        },
    },
    {
        "orig_id": "Enno-Rates-07",
        "robot_file": "RATES-064_staff_view_relationships.robot",
        "env_data": (
            "Staff user: Admin\n"
            "Property: 37 Goldberry Street\n"
            "Relationships: Hammy (Primary), Hilda (Secondary)\n\n"
            "Test navigates to Property record in SF and verifies\n"
            "Property Customer relationships are visible and queryable."
        ),
        "manual_entry": {
            "function": "Staff - Property Management",
            "description": "Staff can view property customer relationships",
            "expected": "Staff can see all customer relationships (Primary, Secondary, Agent) on a property",
            "steps": "1. Navigate to Property record in Salesforce\n2. View Property Customer related list\n3. Verify relationship types and statuses are visible\n4. Confirm via API query",
        },
    },
    {
        "orig_id": "Enno-Rates-08",
        "robot_file": "RATES-064_staff_view_relationships.robot",
        "env_data": (
            "Staff user: Admin\n"
            "Property: 37 Goldberry Street\n\n"
            "Test navigates to Property record, verifies staff can\n"
            "view and query Property Customer relationships."
        ),
        "manual_entry": {
            "function": "Staff - Property Management",
            "description": "Staff can link and unlink property customers",
            "expected": "Staff can create and deactivate property customer relationships",
            "steps": "1. Navigate to Property record in Salesforce\n2. Verify Property Customer records are queryable\n3. Confirm relationships can be managed",
        },
    },
    {
        "orig_id": "MCC-Rates-13",
        "robot_file": "RATES-064_staff_view_relationships.robot",
        "env_data": (
            "Staff user: Admin\n"
            "Property: 37 Goldberry Street\n"
            "Secondary: Hilda Hampster\n\n"
            "Test queries for Secondary relationships and navigates\n"
            "to the property to verify via API."
        ),
        "manual_entry": {
            "function": "Staff - Property Management",
            "description": "Staff can see secondary customers on property",
            "expected": "Staff can identify secondary customer relationships on a property",
            "steps": "1. Navigate to Property in Salesforce\n2. View Property Customer related list\n3. Verify Secondary relationships are visible\n4. Confirm secondary customer name via API",
        },
    },
    {
        "orig_id": "MCC-Rates-14",
        "robot_file": "RATES-065_waive_fees_deflection.robot",
        "env_data": (
            "Person Account: Hammy Hampster\n"
            "Community User: hammy.hampster@test.example.com\n\n"
            "Test checks property summary and payment extension page\n"
            "for waive late fees / interest references and deflection\n"
            "to the request flow."
        ),
        "manual_entry": {
            "function": "Waive Late Fees",
            "description": "Customer is deflected to request flow for fee waiver",
            "expected": "Customer can find the waive late fees / interest option and is directed to submit a request",
            "steps": "1. Log in as Hammy Hampster\n2. Navigate to property summary\n3. Check for waive fees option\n4. Navigate to payment extension page\n5. Verify deflection to request flow",
        },
    },
    {
        "orig_id": "MCC-Rates-19",
        "robot_file": "RATES-066_secondary_deactivates_access.robot",
        "env_data": (
            "Person Account: Hilda Hampster (Secondary)\n"
            "Community User: hilda@scrat.dev1\n"
            "Relationship: Secondary on 37 Goldberry Street\n\n"
            "Test logs in as Hilda, navigates to My Properties & Rates,\n"
            "verifies property is visible, and checks for self-service\n"
            "unlink/remove option."
        ),
        "manual_entry": {
            "function": "Secondary Relationships",
            "description": "Secondary customer can deactivate their own access to a property",
            "expected": "Secondary customer can unlink themselves from a property they no longer wish to manage",
            "steps": "1. Log in as Hilda Hampster (secondary)\n2. Navigate to My Properties & Rates\n3. Verify property is visible\n4. Check for unlink/remove/deactivate option",
        },
    },
]


# CRM-CASE tests (end-to-end case creation tests)
crm_case_tests = [
    {
        "orig_id": "CRM-CASE-01",
        "robot_file": "CRM-CASE-01_short_term_under_5_days.robot",
        "env_data": (
            "Person Account: Hammy Hampster (Primary)\n"
            "Property: 32 Develish Court\n\n"
            "Setup: Quarter dates/balances set dynamically based on current date.\n"
            "Cleanup: Deletes extension case after test.\n\n"
            "Settings: Auto_Approval_Threshold=5, Auto_Approval_Sub_Reason='Under 5 days'"
        ),
        "manual_entry": {
            "function": "Case Creation",
            "description": "Primary Customer can request a short term extension under 5 days",
            "expected": "A primary customer can request a short term arrangement of up to 5 days which is automatically approved and shows on the Property Summary screen",
            "steps": "1. Setup property quarter dates and balances\n2. Login as Hammy Hampster\n3. Navigate to property summary for Develish Court\n4. Click Request payment plan or extension\n5. Click Request a short term arrangement\n6. Verify Up to 5 days is selected\n7. Select Financial Hardship from reason dropdown\n8. Enter additional comments\n9. Click Submit request\n10. Validate confirmation page\n11. Verify case created via API with Approved status\n12. Check arrangement visible on property summary",
        },
    },
    {
        "orig_id": "CRM-CASE-02",
        "robot_file": "CRM-CASE-02_short_term_over_5_days.robot",
        "env_data": (
            "Person Account: Hammy Hampster (Primary)\n"
            "Property: 32 Develish Court\n\n"
            "Setup: Quarter dates/balances set dynamically.\n"
            "Cleanup: Deletes extension case after test.\n\n"
            "Settings: Requires_Approval_Sub_Reason='Over 5 days'"
        ),
        "manual_entry": {
            "function": "Case Creation",
            "description": "Primary Customer can request a short term extension over 5 days",
            "expected": "A primary customer can request a short term arrangement over 5 days which is NOT automatically approved",
            "steps": "1. Setup property and cleanup existing cases\n2. Login as Hammy Hampster\n3. Navigate to property summary\n4. Click Request payment plan or extension\n5. Click Request a short term arrangement\n6. Click More than 5 days\n7. Set a future date\n8. Select Financial Hardship reason\n9. Enter comments\n10. Click Submit\n11. Verify confirmation page\n12. Verify case created with Open status (NOT Approved)\n13. Verify extension link still available on summary",
        },
    },
    {
        "orig_id": "CRM-CASE-03",
        "robot_file": "CRM-CASE-03_staff_approves_arrangement.robot",
        "env_data": (
            "Person Account: Hammy Hampster (Primary)\n"
            "Property: 32 Develish Court\n\n"
            "Test creates an Over 5 days extension (Open status),\n"
            "then admin approves it via API, then customer verifies\n"
            "arrangement shows on summary.\n\n"
            "Cleanup: Deletes case after test."
        ),
        "manual_entry": {
            "function": "Case Creation",
            "description": "Staff Member can approve a Short Term Arrangement Request",
            "expected": "Staff member can approve the short term arrangement and this shows on the Payment Summary Screen",
            "steps": "1. Setup property and cleanup cases\n2. Create Over 5 days extension via portal (Open status)\n3. Verify extension link still available before approval\n4. Admin approves case via API\n5. Login as customer and verify Arrangement visible on summary\n6. Verify Property Requests tab accessible",
        },
    },
    {
        "orig_id": "CRM-CASE-04",
        "robot_file": "CRM-CASE-04_payment_plan_auto_approved.robot",
        "env_data": (
            "Person Account: Hammy Hampster (Primary)\n"
            "Property: 32 Develish Court\n\n"
            "Date entry: Uses 'Keyboard Input type' for LWC date picker.\n"
            "Checkbox: Uses 'span.slds-checkbox_faux' for LWC checkbox.\n"
            "Cleanup: Deletes case after test."
        ),
        "manual_entry": {
            "function": "Case Creation",
            "description": "Primary Customer can request a Payment Plan (auto-approved)",
            "expected": "A primary customer can request a payment plan which is automatically approved and shows on the Property Summary screen",
            "steps": "1. Setup property and cleanup cases\n2. Login as Hammy Hampster\n3. Navigate to property summary\n4. Click Request payment plan or extension\n5. Click Request a payment plan\n6. Enter start date\n7. Select Weekly frequency\n8. Scroll and click Continue\n9. Accept terms and conditions\n10. Click Request plan\n11. Verify confirmation page\n12. Verify case created via API\n13. Admin sets plan to Active\n14. Verify Weekly Payment Plan on summary",
        },
    },
    {
        "orig_id": "CRM-CASE-05",
        "robot_file": "CRM-CASE-05_payment_plan_needs_approval.robot",
        "env_data": (
            "Person Account: Hammy Hampster (Primary)\n"
            "Property: 32 Develish Court\n\n"
            "Sets Allow_Payment_Plan_Auto_Approval=No before test.\n"
            "Restores original setting after test.\n"
            "Cleanup: Deletes case after test."
        ),
        "manual_entry": {
            "function": "Case Creation",
            "description": "Primary Customer can request a payment plan that needs approval",
            "expected": "A primary customer can request a payment plan that is NOT automatically approved. Staff approves and plan shows on summary.",
            "steps": "1. Set Allow_Payment_Plan_Auto_Approval=No\n2. Setup property and cleanup cases\n3. Login and complete payment plan form\n4. Verify case created\n5. Admin approves and sets Active\n6. Verify Weekly Payment Plan on summary\n7. Check Property Requests tab\n8. Restore setting and cleanup",
        },
    },
    {
        "orig_id": "CRM-CASE-08",
        "robot_file": "CRM-CASE-08_transfer_payment.robot",
        "env_data": (
            "Person Account: Hammy Hampster (Primary)\n"
            "Property: 32 Develish Court\n\n"
            "Pre-requisite: Payment history API must return records.\n"
            "Waits up to 60s for payment data to load.\n"
            "Cleanup: Deletes transfer case after test."
        ),
        "manual_entry": {
            "function": "Case Creation",
            "description": "Primary customer transfers a payment from one property to another",
            "expected": "Customer can select a payment, request transfer to another linked property, and a transfer case is created",
            "steps": "1. Setup property and cleanup cases\n2. Login as Hammy Hampster\n3. Navigate to Payments tab\n4. Wait for payment records to load from API\n5. Click Actions dropdown on first payment\n6. Click Request a refund or transfer\n7. Click Transfer to another linked property\n8. Verify transfer amount displayed\n9. Select target property (Goldberry Street)\n10. Click Submit Request\n11. Verify confirmation page\n12. Verify transfer case created via API",
        },
    },
    {
        "orig_id": "CRM-CASE-09",
        "robot_file": "CRM-CASE-09_refund_payment.robot",
        "env_data": (
            "Person Account: Hammy Hampster (Primary)\n"
            "Property: 32 Develish Court\n\n"
            "Pre-requisite: Payment history API must return records.\n"
            "Waits up to 60s for payment data to load.\n"
            "Cleanup: Deletes reverse payment case after test."
        ),
        "manual_entry": {
            "function": "Case Creation",
            "description": "Primary customer can refund a payment made in error",
            "expected": "Customer can select a payment, request refund back to them, and a reverse payment case is created",
            "steps": "1. Setup property and cleanup cases\n2. Login as Hammy Hampster\n3. Navigate to Payments tab\n4. Wait for payment records to load\n5. Click Actions dropdown\n6. Click Request a refund or transfer\n7. Click Refund back to me\n8. Verify refund amount displayed\n9. Check upload documents section\n10. Click Submit Request\n11. Verify confirmation page\n12. Verify Reverse Payment case created via API",
        },
    },
    {
        "orig_id": "CRM-CASE-10",
        "robot_file": "CRM-CASE-10_refund_credit_balance.robot",
        "env_data": (
            "Person Account: Hammy Hampster (Primary)\n"
            "Property: 37 Goldberry Street (credit balance -$300)\n\n"
            "Setup: Sets Full_year_balance to -300 (credit).\n"
            "BSB: 035099, Account: 100000\n"
            "Cleanup: Deletes refund credit case after test."
        ),
        "manual_entry": {
            "function": "Case Creation",
            "description": "Customer can refund a credit balance with BSB details",
            "expected": "Customer with credit balance can request refund, enter BSB/account details, and a Refund Credit Balance case is created",
            "steps": "1. Setup property with credit balance (-$300)\n2. Login as Hammy Hampster\n3. Navigate to Goldberry Street summary\n4. Click Refund credit link\n5. Select full credit balance\n6. Enter BSB (035099) and Account Number (100000)\n7. Click Submit Request\n8. Verify confirmation with reference number\n9. Check Property Requests tab\n10. Verify Refund Credit Balance case created via API",
        },
    },
]


def main():
    new_wb = openpyxl.load_workbook("C:/Users/hughw/OneDrive/Documents/ennovative/ProjectsNewLaptop/CFSuiteFullBuildUnpackaged/CFSuiteCumulusFullBuildUnpackaged/documentation/Automated Testing.xlsx")
    ws = new_wb["Automated Tests"]

    wrap = Alignment(wrap_text=True, vertical="top")

    current_row = 4

    # Write setup tests
    for t in setup_tests:
        ws.cell(row=current_row, column=1, value=t["function"]).alignment = wrap
        ws.cell(row=current_row, column=2, value=t["id"]).alignment = wrap
        ws.cell(row=current_row, column=3, value=t["description"]).alignment = wrap
        ws.cell(row=current_row, column=4, value=t["expected"]).alignment = wrap
        ws.cell(row=current_row, column=5, value=t["steps"]).alignment = wrap
        ws.cell(row=current_row, column=6, value=t["env_data"]).alignment = wrap
        current_row += 1

    # Write implemented tests (Account Establishment + Property Details)
    all_tests = (implemented_tests + new_rates_cases_tests + payment_history_tests
                 + payment_plan_tests + property_details_tests + rates_notices_tests
                 + making_payments_tests + refund_transfer_tests + ept_tests
                 + crm_case_tests)
    for test in all_tests:
        orig_id = test["orig_id"]
        orig = original_tests.get(orig_id, {})
        manual = test.get("manual_entry", {})

        function = manual.get("function") or orig.get("function", "Account Establishment")
        description = manual.get("description") or orig.get("description", "")
        expected = manual.get("expected") or orig.get("expected", "")
        steps = manual.get("steps") or orig.get("steps", "")

        # Only use our test environment data (not original client test data)
        env_data = test.get("env_data", "")
        parts = []
        if env_data:
            parts.append(f"Test Environment (dev_namespaced_1):\n{env_data}")
        parts.append(f"Robot File: {test['robot_file']}")
        combined_data = "\n\n".join(parts)

        ws.cell(row=current_row, column=1, value=function).alignment = wrap
        ws.cell(row=current_row, column=2, value=orig_id).alignment = wrap
        ws.cell(row=current_row, column=3, value=description).alignment = wrap
        ws.cell(row=current_row, column=4, value=expected).alignment = wrap
        ws.cell(row=current_row, column=5, value=steps).alignment = wrap
        ws.cell(row=current_row, column=6, value=combined_data).alignment = wrap
        current_row += 1

    # Set column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 55
    ws.column_dimensions["F"].width = 55

    new_wb.save("C:/Users/hughw/OneDrive/Documents/ennovative/ProjectsNewLaptop/CFSuiteFullBuildUnpackaged/CFSuiteCumulusFullBuildUnpackaged/documentation/Automated Testing.xlsx")
    print(f"Spreadsheet updated: {current_row - 4} rows (rows 4-{current_row - 1})")
    print("Saved to: C:/Users/hughw/OneDrive/Documents/ennovative/ProjectsNewLaptop/CFSuiteFullBuildUnpackaged/CFSuiteCumulusFullBuildUnpackaged/documentation/Automated Testing.xlsx")


if __name__ == "__main__":
    main()
